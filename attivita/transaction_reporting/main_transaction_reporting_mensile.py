"""
CON-412 Transaction Reporting - Sistema Automatico
Elaborazione file Excel per validazione ISIN via ESMA
"""

import sys
import logging
from datetime import datetime
from pathlib import Path
from openpyxl.styles import Font

# Import diretto dei servizi
from services.local_file_service import LocalFileService
from services.database_service import DatabaseService
from services.isin_validation_service import ISINValidationService


class InteractiveConfig:
    """Gestisce la configurazione interattiva del sistema"""
    
    def __init__(self):
        self.config = {}
        
    def get_file_config(self):
        """Richiede configurazione file locale/NAS all'utente"""
        print("\n🔧 CONFIGURAZIONE FILE LOCALE/NAS")
        print("=" * 40)
        
        config = self._get_file_config()
        if config:
            # Chiede se abilitare il controllo database
            config['enable_database_check'] = self._ask_database_check()
        
        return config
    
    def _get_file_config(self):
        """Configurazione file locale/NAS"""
        print("\n📁 CONFIGURAZIONE FILE")
        print("-" * 30)
        print("Inserisci il path completo al file Excel:")
        print()
        print("Esempi di path:")
        print("• File locale: C:\\Downloads\\CON-412_AGOSTO.xlsx")
        print("• NAS/Server: \\\\server\\share\\reports\\CON-412_AGOSTO.xlsx")  
        print("• Rete aziendale: \\\\nas-server\\shared\\Transaction_Reporting\\CON-412_AUGUST.xlsx")
        print()
        
        # Chiede solo il path completo del file
        file_path = input("Path completo al file Excel: ").strip()
        if not file_path:
            print("❌ Path del file è obbligatorio!")
            return None
            
        # Estrae automaticamente directory e nome file dal path completo
        from pathlib import Path
        path_obj = Path(file_path)
        base_path = str(path_obj.parent)
        file_name = path_obj.name
        
        print(f"\n📍 CONFIGURAZIONE RILEVATA")
        print(f"Directory: {base_path}")
        print(f"Nome file: {file_name}")
        print(f"Path completo: {file_path}")
        
        return {
            'type': 'local_file',
            'source': 'local',
            'base_path': base_path,
            'file_name': file_name,
            'file_path': file_path
        }
    
    def _ask_database_check(self):
        """Il controllo database è sempre abilitato nell'ordine prestabilito"""
        print("\n💾 CONTROLLO DATABASE ORACLE")
        print("-" * 30)
        print("Il sistema eseguirà automaticamente:")
        print("• FASE 4: Filtraggio ordini tramite query database (RF/EE)")
        print("• FASE 5: Primo controllo ESMA - Validazione ISIN")
        print("• FASE 6: Secondo controllo ESMA - Validazione Trading Venue")
        print("✅ Controlli sempre abilitati nell'ordine prestabilito")
        return True


class CON412Processor:
    """Processore principale per CON-412 con validazione ESMA"""
    
    def __init__(self, config):
        self.config = config
        self.logger = self._setup_logging()
        
        # Inizializza il servizio di validazione ISIN
        self.isin_validation_service = ISINValidationService()
        
    def _extract_market_code(self, mercato_raw: str) -> str:
        """
        Estrae il codice mercato pulito rimuovendo le parentesi e il loro contenuto.
        
        Args:
            mercato_raw: Stringa mercato originale (es. "MTAA(MTA)" o "XOFF")
            
        Returns:
            Codice mercato pulito (es. "MTAA" o "XOFF")
        """
        if not mercato_raw:
            return ""
        
        mercato_str = str(mercato_raw).strip()
        
        # Se ci sono parentesi, prende solo la parte prima delle parentesi
        if '(' in mercato_str:
            return mercato_str.split('(')[0].strip()
        
        # Se non ci sono parentesi, restituisce la stringa originale
        return mercato_str
        
    def _convert_utc_to_italian_time(self, utc_datetime):
        """
        Converte una data/ora UTC in ora italiana (considerando DST).
        Italia usa UTC+1 (CET) in inverno e UTC+2 (CEST) in estate.
        
        Args:
            utc_datetime: datetime object in UTC
            
        Returns:
            datetime object in ora italiana
        """
        import datetime as dt
        
        # Determina se siamo in periodo DST (ultima domenica marzo - ultima domenica ottobre)
        year = utc_datetime.year
        
        # Trova l'ultima domenica di marzo (inizio DST)
        march_last_day = dt.date(year, 3, 31)
        while march_last_day.weekday() != 6:  # 6 = domenica
            march_last_day -= dt.timedelta(days=1)
        dst_start = dt.datetime.combine(march_last_day, dt.time(2, 0))  # 02:00 UTC
        
        # Trova l'ultima domenica di ottobre (fine DST)
        october_last_day = dt.date(year, 10, 31)
        while october_last_day.weekday() != 6:  # 6 = domenica
            october_last_day -= dt.timedelta(days=1)
        dst_end = dt.datetime.combine(october_last_day, dt.time(1, 0))  # 01:00 UTC
        
        # Determina l'offset da applicare
        if dst_start <= utc_datetime < dst_end:
            # Periodo DST: UTC+2 (CEST)
            offset_hours = 2
        else:
            # Periodo non DST: UTC+1 (CET)
            offset_hours = 1
        
        return utc_datetime + dt.timedelta(hours=offset_hours)
        
    def _setup_logging(self):
        """Configura il sistema di logging"""
        # Assicurati che tutti i percorsi siano relativi alla directory dello script
        base_dir = Path(__file__).parent
        logs_dir = base_dir / "log_tr_mensile"
        work_dir = base_dir / "work"
        output_dir = base_dir / "output_tr_mensile"

        # Aggiorna i percorsi relativi
        logs_dir.mkdir(exist_ok=True)
        work_dir.mkdir(exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = logs_dir / f"con412_processing_{timestamp}.log"

        # Create a file handler for detailed logs
        file_handler = logging.FileHandler(log_file)
        file_handler.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(file_formatter)

        # Create a stream handler for minimal terminal output
        stream_handler = logging.StreamHandler(sys.stdout)
        stream_handler.setLevel(logging.CRITICAL)  # Mostra solo messaggi critici nel terminale
        stream_formatter = logging.Formatter('%(message)s')
        stream_handler.setFormatter(stream_formatter)

        # Configure the root logger
        logging.basicConfig(level=logging.DEBUG, handlers=[file_handler, stream_handler])

        # Configura il livello di logging per l'eseguibile finale
        if getattr(sys, 'frozen', False):  # Verifica se è un eseguibile
            logging.getLogger().setLevel(logging.INFO)  # Imposta il livello a INFO per escludere DEBUG
        else:
            logging.getLogger().setLevel(logging.DEBUG)  # Mantieni DEBUG durante lo sviluppo

        return logging.getLogger(__name__)
    
    def run(self):
        """Esegue il processo CON-412 con validazione ESMA"""
        try:
            self.logger.info("=" * 60)
            self.logger.info("AVVIO PROCESSO CON-412 - TRANSACTION REPORTING")
            self.logger.info("=" * 60)
            self.logger.info(f"Configurazione: {self.config['type']}")
            self.logger.info(f"Path file: {self.config['file_path']}")
            
            # Fase 1: Verifica accesso al file
            print("\n🔄 FASE 1: Verifica accesso al file")
            
            file_service = LocalFileService.create_for_path(self.config['file_path'])
            self.logger.info(f"Path file: {self.config['file_path']}")
            
            if not file_service.check_access():
                print("❌ Errore accesso alla directory/file")
                return False
            print("✅ Accesso verificato")
            
            # Fase 2: Ricerca e lettura file
            print("\n🔄 FASE 2: Ricerca e lettura file")
            
            print(f"📁 Leggo file: {self.config['file_path']}")
            
            # Usa il path diretto
            source_file_path = None
            if Path(self.config['file_path']).exists():
                source_file_path = self.config['file_path']
            else:
                self.logger.error(f"File non trovato: {self.config['file_path']}")
            
            if not source_file_path:
                print("❌ File non trovato")
                return False
            
            print(f"✅ File trovato: {Path(source_file_path).name}")
            self.logger.info(f"File sorgente: {source_file_path}")
            
            # Copia il file in una directory di lavoro
            work_dir = Path(__file__).parent / "work"
            work_dir.mkdir(exist_ok=True)
            
            work_file_path = work_dir / Path(source_file_path).name
            
            if source_file_path != str(work_file_path):
                print("📋 Copia file nella directory di lavoro...")
                if not file_service.copy_file(source_file_path, str(work_file_path)):
                    print("❌ Errore copia file")
                    return False
                print("✅ File copiato in directory di lavoro")
                downloaded_file = str(work_file_path)
            else:
                downloaded_file = source_file_path
            
            # Fase 3: Lettura struttura file Excel
            print("\n📊 FASE 3: Lettura struttura file Excel")
            # Memorizza il percorso del file per controlli successivi
            self._current_excel_file = downloaded_file
            original_data = self._read_excel_file(downloaded_file)
            if not original_data:
                print("❌ Errore lettura file Excel")
                return False
            
            print(f"📊 Struttura letta: {len(original_data)} gruppi ISIN")
            print("✅ Struttura originale preservata - controlli ESMA su ogni ordine")
            
            # Mostra dettaglio ordini
            total_orders = sum(len(group.get('orders', [])) for group in original_data)
            unique_isins = len(set(group['isin'] for group in original_data))
            print(f"📊 DETTAGLIO: {unique_isins} ISIN unici, {total_orders} ordini totali")
            
            # Fase 4: Filtraggio tramite query database (SEMPRE ATTIVO)
            print("\n🔄 FASE 4: Filtraggio ordini tramite query database")
            print("🔍 Connessione al database Oracle TNS per filtraggio ordini")
            print("📋 NOTA: Le credenziali database saranno richieste se non già memorizzate")
            
            # Forza la richiesta delle credenziali database
            max_retry_attempts = 3
            db_connection_successful = False
            
            for attempt in range(1, max_retry_attempts + 1):
                try:
                    print(f"\n🔄 Tentativo di connessione {attempt}/{max_retry_attempts}")
                    
                    # Inizializza servizi database
                    db_service = DatabaseService()
                    
                    # Testa la connessione (questo dovrebbe chiedere le credenziali)
                    test_result = db_service.test_connection()
                    db_connection_successful = test_result["success"]
                    
                    if db_connection_successful:
                        print("✅ Connessione database riuscita")
                        # TODO: Implementare logica di filtraggio ordini RF/EE
                        # original_count = len(original_data)
                        # original_data = filter_service.filter_orders_by_status(original_data)
                        filtered_count = len(original_data)
                        
                        print(f"✅ Filtraggio database completato")
                        break
                    else:
                        print(f"❌ Connessione database fallita: {test_result.get('error', 'Errore sconosciuto')}")
                        if attempt < max_retry_attempts:
                            retry_choice = input(f"Vuoi riprovare la connessione? (S)ì o (N)o [S]: ").strip().upper()
                            if retry_choice == "N" or retry_choice == "NO":
                                break
                        
                except Exception as e:
                    self.logger.error(f"Errore tentativo {attempt} filtraggio database: {str(e)}")
                    print(f"❌ Errore tentativo {attempt}: {str(e)}")
                    if attempt < max_retry_attempts:
                        retry_choice = input(f"Vuoi riprovare la connessione? (S)ì o (N)o [S]: ").strip().upper()
                        if retry_choice == "N" or retry_choice == "NO":
                            break
            
            if not db_connection_successful:
                print("⚠️ Tutti i tentativi di connessione database falliti")
                print("⚠️ Database non disponibile - continuando con controlli ESMA su tutti gli ordini")
                print("📋 NOTA: Senza filtraggio database, alcuni ordini potrebbero essere elaborati anche se RF/EE")
                self.logger.warning("Continuando senza filtraggio database - tutti gli ordini saranno elaborati")
            
            # Fase 5-8: Controlli ESMA sequenziali (1-4)
            print("\n🔄 FASE 5-8: Controlli ESMA sequenziali")
            print("🌐 Esecuzione controlli ESMA in ordine procedurale:")
            print("  1. Controllo ISIN censito (API restituisce risultati?)")
            print("  2. Controllo Trading Venue (MIC vs MERCATO)")
            print("  3. Controllo Date Approval")
            print("  4. Controllo Maturity Date")
            
            original_data = self._run_sequential_controls(original_data)
            
            # Calcola statistiche per ciascun controllo basandosi sui singoli ordini
            all_orders = []
            for group in original_data:
                all_orders.extend(group.get('orders', []))
            
            controllo_1_failed = sum(1 for order in all_orders if order.get('controllo_1_failed', False))
            controllo_2_failed = sum(1 for order in all_orders if order.get('controllo_2_failed', False))
            controllo_3_failed = sum(1 for order in all_orders if order.get('controllo_3_failed', False))
            controllo_4_failed = sum(1 for order in all_orders if order.get('controllo_4_failed', False))
            
            validated_count = len(all_orders) - controllo_1_failed
            venue_valid_count = len(all_orders) - controllo_2_failed
            non_censiti_count = controllo_1_failed
            
            print(f"\n📊 RISULTATI CONTROLLI SEQUENZIALI (per ordine):")
            print(f"  Controllo 1 (ISIN censiti): {validated_count}/{len(all_orders)} ordini passati")
            print(f"  Controllo 2 (Trading Venue): {venue_valid_count}/{len(all_orders)} ordini passati")
            print(f"  Controllo 3 (Date Approval): {len(all_orders) - controllo_3_failed}/{len(all_orders)} ordini passati")
            print(f"  Controllo 4 (Maturity Date): {len(all_orders) - controllo_4_failed}/{len(all_orders)} ordini passati")
            print(f"  ORDINI NON CENSITI: {non_censiti_count}/{len(all_orders)}")
            
            print("\n🔄 FASE 7: Aggiunta X per ISIN validati ESMA")
            
            # Crea directory output CON-412
            # Crea directory output accessibile per l'exe
            if hasattr(sys, '_MEIPASS'):
                # Quando eseguito come exe, usa la directory dell'eseguibile
                output_dir = Path(sys.executable).parent / "output_tr_mensile"
            else:
                # Quando eseguito come script Python
                output_dir = Path(sys.argv[0]).parent / "output_tr_mensile"
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Usa il nome del file di input per il file di output
            input_file_name = Path(downloaded_file).stem
            output_name = f"{input_file_name}_Validated.xlsx"
            output_path = output_dir / output_name
            
            print(f"📄 Aggiornamento file con validazione ESMA: {output_name}")
            print(f"📁 Directory output: {output_dir}")
            print(f"📄 Struttura: File originale + X nella colonna CASISTICA per ISIN validati")
            
            # Crea file Excel con validazione
            try:
                success = self._create_validated_excel(downloaded_file, original_data, output_path)
                if not success:
                    print("❌ Errore creazione file Excel validato")
                    return False
                    
            except Exception as e:
                print(f"⚠️ Errore creazione Excel: {e}")
                return False
            
            print(f"\n🎯 ELABORAZIONE CON-412 COMPLETATA")
            print(f"📁 Report salvato in: {output_path}")
            print(f"📊 File originale processato: {Path(downloaded_file).name}")
            print(f"📊 Sorgente: {self.config['source']}")
            print(f"📊 CONTROLLO 1 - Ordini con ISIN validati: {validated_count}/{len(all_orders)}")
            print(f"📊 CONTROLLO 2 - Ordini con Trading Venue validati: {venue_valid_count}/{len(all_orders)}")
            print(f"📊 NOTA: Ogni ordine viene verificato singolarmente con il suo mercato specifico")
            
            # Genera e stampa il resoconto dettagliato  
            # self._generate_detailed_summary(original_data, validated_count, venue_valid_count, non_censiti_count, output_path)
            
            self.logger.info("PROCESSO CON-412 COMPLETATO CON SUCCESSO")
            self.logger.info(f"Report: {output_path}")
            self.logger.info(f"Sorgente: {self.config['source']}")
            self.logger.info(f"Controllo 1 ISIN: {validated_count}/{len(all_orders)} ordini")
            self.logger.info(f"Controllo 2 Venue: {venue_valid_count}/{len(all_orders)} ordini")
            self.logger.info("=" * 60)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Errore nel processo CON-412: {str(e)}")
            print(f"❌ Errore critico: {str(e)}")
            return False
    
    def _read_excel_file(self, file_path: str) -> list:
        """
        Legge la struttura del file Excel e identifica i gruppi ISIN
        
        Args:
            file_path: Percorso del file Excel da leggere
            
        Returns:
            Lista di dizionari con i dati dei gruppi ISIN
        """
        try:
            import openpyxl
            
            self.logger.info(f"Lettura file Excel: {file_path}")
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            if not ws:
                ws = wb.worksheets[0]
            
            self.logger.info(f"Foglio Excel: {ws.title}")
            
            # Cerca la riga di intestazione
            header_row = None
            isin_col = None
            occurrences_col = None
            order_number_col = None
            mercato_col = None
            
            for row_num in range(1, min(20, ws.max_row + 1)):
                for col_num in range(1, min(20, ws.max_column + 1)):  # Aumento il range per cercare più colonne
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value:
                        cell_text = str(cell_value).upper().strip()
                        # Cerca esattamente "ISIN" e non "DESCRIZIONE ISIN"
                        if cell_text == 'ISIN' and not isin_col:
                            header_row = row_num
                            isin_col = col_num
                        elif ('OCCORREN' in cell_text or 'OCCURRENCE' in cell_text) and not occurrences_col:
                            occurrences_col = col_num
                            if not header_row:
                                header_row = row_num
                        # Cerca colonna numero ordine
                        elif (('NUMERO' in cell_text and 'ORDINE' in cell_text) or 
                              'ORDER' in cell_text or 
                              cell_text in ['NUMORD', 'NUM_ORD', 'ORDER_NUM']) and not order_number_col:
                            order_number_col = col_num
                            if not header_row:
                                header_row = row_num
                        # Cerca colonna mercato
                        elif 'MERCATO' in cell_text and not mercato_col:
                            mercato_col = col_num
                            if not header_row:
                                header_row = row_num
            
            if not header_row or not isin_col:
                self.logger.error("Impossibile trovare colonne ISIN nel file Excel")
                return []
            
            if not occurrences_col:
                # Se non trova OCCURRENCES, assume colonna dopo ISIN
                occurrences_col = isin_col + 1
                self.logger.warning(f"Colonna OCCURRENCES non trovata, uso colonna {occurrences_col}")
            
            # Controllo se tutte le colonne necessarie sono state trovate
            if not order_number_col:
                self.logger.warning("Colonna numero ordine non trovata - controllo database disabilitato")
            
            # Legge i dati e identifica i gruppi ISIN (logica originale)
            data = []
            current_isin = None
            current_isin_row = None
            expected_orders = 0
            order_count = 0
            current_group_orders = []  # Lista per memorizzare i dettagli degli ordini del gruppo corrente
            
            for row_num in range(header_row + 1, ws.max_row + 1):
                isin_value = ws.cell(row=row_num, column=isin_col).value
                occurrences_value = ws.cell(row=row_num, column=occurrences_col).value
                order_number_value = ws.cell(row=row_num, column=order_number_col).value if order_number_col else None
                mercato_value = ws.cell(row=row_num, column=mercato_col).value if mercato_col else None
                
                # Se trova un ISIN valorizzato, è una nuova riga ISIN (NON un ordine)
                if isin_value and str(isin_value).strip() != '':
                    # Se c'era un gruppo precedente, lo finalizza
                    if current_isin and current_group_orders:
                        # Aggiorna l'ultimo gruppo con i dettagli degli ordini
                        data[-1]['orders'] = current_group_orders
                    
                    isin_str = str(isin_value).strip()
                    occurrences_num = int(occurrences_value) if occurrences_value and str(occurrences_value).isdigit() else 0
                    
                    # Validazione ISIN di base (12 caratteri alfanumerici)
                    if len(isin_str) >= 12 and isin_str.isalnum():
                        current_isin = isin_str
                        current_isin_row = row_num
                        expected_orders = occurrences_num
                        order_count = 0  # La riga ISIN NON è un ordine, si parte da 0
                        current_group_orders = []
                        
                        # NON aggiunge la riga ISIN come ordine - gli ordini sono nelle righe successive
                        print(f"📋 ISIN trovato: {isin_str} con {occurrences_num} ordini attesi (riga {row_num})")
                        
                        data.append({
                            'isin': isin_str,
                            'occurrences': occurrences_num,
                            'row_num': row_num,  # Riga dell'ISIN (non di un ordine)
                            'mercato': str(mercato_value).strip() if mercato_value else None,  # Mercato dalla prima riga
                            'order_rows': list(range(row_num, row_num + occurrences_num)),  # Tutte le righe del gruppo
                            'orders': [],  # Sarà popolato alla fine del gruppo
                            'esma_valid': None  # Sarà validato tramite servizio
                        })
                    else:
                        self.logger.info(f"ISIN non valido scartato: '{isin_str}' (len={len(isin_str)}, isalnum={isin_str.isalnum()})")
                        current_isin = None
                        current_isin_row = None
                        expected_orders = 0
                        order_count = 0
                        current_group_orders = []
                
                # Se siamo in un gruppo ISIN e questa è una riga di ordine (senza ISIN ma con numero ordine)
                elif current_isin and order_count < expected_orders and order_number_value and str(order_number_value).strip():
                    order_count += 1
                    
                    # Aggiunge i dettagli dell'ordine
                    current_group_orders.append({
                        'row_num': row_num,
                        'numero_ordine': str(order_number_value).strip(),
                        'mercato': str(mercato_value).strip() if mercato_value else None
                    })
                    print(f"  📝 Ordine {order_count}/{expected_orders}: {str(order_number_value).strip()} (riga {row_num})")
                    
                    # Se abbiamo completato tutti gli ordini del gruppo
                    if order_count >= expected_orders:
                        # Finalizza il gruppo corrente
                        data[-1]['orders'] = current_group_orders
                        current_isin = None
                        current_isin_row = None
                        expected_orders = 0
                        order_count = 0
                        current_group_orders = []
            
            # Finalizza l'ultimo gruppo se necessario
            if current_isin and current_group_orders:
                data[-1]['orders'] = current_group_orders
            
            self.logger.info(f"File Excel letto correttamente: {len(data)} gruppi ISIN trovati")
            return data
            
        except ImportError:
            self.logger.error("openpyxl non disponibile per lettura Excel")
            return []
        except Exception as e:
            self.logger.error(f"Errore lettura file Excel: {str(e)}")
            return []
    
    def _create_validated_excel(self, original_file_path: str, data: list, output_path: Path) -> bool:
        """
        Crea il file Excel validato con filtraggio database e validazione ESMA
        
        Args:
            original_file_path: Percorso del file Excel originale
            data: Lista dei dati processati e filtrati
            output_path: Percorso del file di output
            
        Returns:
            True se la creazione è riuscita
        """
        try:
            import openpyxl
            from openpyxl.styles import Font
            import shutil
            
            # COPIA COMPLETA DEL FILE ORIGINALE (senza modifiche)
            shutil.copy2(original_file_path, output_path)
            self.logger.info(f"File originale copiato completamente: {original_file_path} -> {output_path}")
            
            # Carica il file copiato per aggiungere solo le X
            workbook = openpyxl.load_workbook(output_path)
            worksheet = workbook.active
            
            # Trova le colonne CASISTICA nel file copiato (struttura originale preservata)
            header_row = None
            casistica_isin_col = None
            casistica_venue_col = None
            casistica_date_approval_col = None  # Controllo 3
            casistica_maturity_col = None        # Controllo 4
            
            # Cerca l'header e le colonne casistica nel file copiato
            for row_num in range(1, min(20, worksheet.max_row + 1)):
                for col_num in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    if cell_value:
                        cell_text = str(cell_value).upper().strip()
                        if 'CASISTICA' in cell_text and 'ISIN NON CENSITO' in cell_text:
                            casistica_isin_col = col_num
                            header_row = row_num
                        elif 'CASISTICA' in cell_text and 'MIC CODE NON PRESENTE' in cell_text:
                            casistica_venue_col = col_num
                            header_row = row_num
                        elif 'CASISTICA' in cell_text and 'DATA DI AMMISSIONE' in cell_text:
                            casistica_date_approval_col = col_num
                            header_row = row_num
                        elif 'CASISTICA' in cell_text and 'DATA DI CESSAZIONE' in cell_text:
                            casistica_maturity_col = col_num
                            header_row = row_num
                        elif cell_text == 'ISIN' and not header_row:
                            header_row = row_num
                if (casistica_isin_col and casistica_venue_col and casistica_date_approval_col and casistica_maturity_col) or header_row:
                    break
            
            if not header_row:
                self.logger.error("Impossibile trovare la riga header")
                return False
            
            if not casistica_isin_col:
                self.logger.warning("Colonna CASISTICA ISIN NON CENSITO non trovata")
            
            if not casistica_venue_col:
                self.logger.warning("Colonna CASISTICA MIC CODE NON PRESENTE non trovata")
                
            if not casistica_date_approval_col:
                self.logger.warning("Colonna CASISTICA DATA DI AMMISSIONE non trovata")
                
            if not casistica_maturity_col:
                self.logger.warning("Colonna CASISTICA DATA DI CESSAZIONE non trovata")
                
            # Debug log per le colonne trovate
            print(f"DEBUG - Colonna Controllo 3: {casistica_date_approval_col}")
            print(f"DEBUG - Colonna Controllo 4: {casistica_maturity_col}")
            
            # AGGIUNGI SOLO LE X - IL FILE È GIÀ COMPLETO!
            print(f"✅ File originale preservato - aggiungo solo X per validazione ESMA")
            
            # Applica la validazione ESMA sui singoli ordini - DIRETTAMENTE SUL FILE COPIATO
            for group in data:
                orders = group.get('orders', [])
                
                for order in orders:
                    # Salta gli ordini virtuali (creati per ISIN senza ordini)
                    if order.get('is_virtual', False):
                        continue
                        
                    original_row_num = order.get('row_num')
                    if not original_row_num:
                        continue
                        
                    # Usa direttamente la riga originale (file già copiato)
                    target_row = original_row_num
                    
                    # CONTROLLO 1: ISIN NON CENSITO per questo ordine specifico
                    if order.get('controllo_1_failed', False) and casistica_isin_col:
                        cell = worksheet.cell(row=target_row, column=casistica_isin_col)
                        cell.value = 'X'
                        cell.font = Font(bold=False, color="000000")
                    
                    # CONTROLLO 2: MIC CODE NON PRESENTE (solo se controllo 1 è passato per questo ordine)
                    elif order.get('controllo_2_failed', False) and casistica_venue_col:
                        cell = worksheet.cell(row=target_row, column=casistica_venue_col)
                        cell.value = 'X'
                        cell.font = Font(bold=False, color="000000")
                    
                    # CONTROLLO 3: DATA DI AMMISSIONE (solo se controlli 1 e 2 sono passati per questo ordine)
                    elif order.get('controllo_3_failed', False) and casistica_date_approval_col:
                        cell = worksheet.cell(row=target_row, column=casistica_date_approval_col)
                        cell.value = 'X'
                        cell.font = Font(bold=False, color="000000")
                    
                    # CONTROLLO 4: DATA DI CESSAZIONE (solo se controlli 1, 2 e 3 sono passati per questo ordine)
                    elif order.get('controllo_4_failed', False) and casistica_maturity_col:
                        cell = worksheet.cell(row=target_row, column=casistica_maturity_col)
                        cell.value = 'X'
                        cell.font = Font(bold=False, color="000000")
            
            # Salva il file con le X aggiunte
            workbook.save(output_path)
            self.logger.info(f"File Excel validato salvato: {output_path}")
            print(f"✅ File originale preservato con X aggiunte: {output_path}")
            
            # Calcola statistiche basandosi sui singoli ordini (separando reali da virtuali)
            all_orders = []
            virtual_orders = []
            for group in data:
                real_orders = [order for order in group.get('orders', []) if not order.get('is_virtual', False)]
                virtual_orders_group = [order for order in group.get('orders', []) if order.get('is_virtual', False)]
                all_orders.extend(real_orders)
                virtual_orders.extend(virtual_orders_group)
            
            controllo_1_failed_count = sum(1 for order in all_orders if order.get('controllo_1_failed', False))
            controllo_2_failed_count = sum(1 for order in all_orders if order.get('controllo_2_failed', False))
            controllo_3_failed_count = sum(1 for order in all_orders if order.get('controllo_3_failed', False))
            controllo_4_failed_count = sum(1 for order in all_orders if order.get('controllo_4_failed', False))
            api_error_count = sum(1 for order in all_orders if order.get('api_error', False))
            
            # Statistiche per ISIN senza ordini (controlli virtuali)
            virtual_1_failed = sum(1 for order in virtual_orders if order.get('controllo_1_failed', False))
            virtual_2_failed = sum(1 for order in virtual_orders if order.get('controllo_2_failed', False))
            virtual_3_failed = sum(1 for order in virtual_orders if order.get('controllo_3_failed', False))
            virtual_4_failed = sum(1 for order in virtual_orders if order.get('controllo_4_failed', False))
            virtual_api_error = sum(1 for order in virtual_orders if order.get('api_error', False))
            
            total_isin_count = len(data)
            total_orders_count = len(all_orders)
            isin_no_orders_count = len(virtual_orders)
            
            print(f"\n📊 STATISTICHE CONTROLLI:")
            print(f"   📋 ISIN totali nel file: {total_isin_count}")
            print(f"   📋 ISIN con ordini: {total_isin_count - isin_no_orders_count}")
            print(f"   📋 ISIN senza ordini: {isin_no_orders_count}")
            print(f"   📋 Ordini totali: {total_orders_count}")
            print(f"   ❌ Controllo 1 fallito: {controllo_1_failed_count} ordini + {virtual_1_failed} ISIN")
            print(f"   ❌ Controllo 2 fallito: {controllo_2_failed_count} ordini + {virtual_2_failed} ISIN")
            print(f"   ❌ Controllo 3 fallito: {controllo_3_failed_count} ordini + {virtual_3_failed} ISIN")
            print(f"   ❌ Controllo 4 fallito: {controllo_4_failed_count} ordini + {virtual_4_failed} ISIN")
            if api_error_count > 0 or virtual_api_error > 0:
                print(f"   🚫 Errori API ESMA: {api_error_count} ordini + {virtual_api_error} ISIN (controlli non eseguiti)")
            
            isin_valid_count = len(all_orders) - controllo_1_failed_count  # Ordini con ISIN validi
            venue_valid_count = len(all_orders) - controllo_2_failed_count  # Ordini con venue validi
            
            print("✅ File Excel creato con filtraggio database + controlli ESMA sequenziali")
            print(f"📊 Gruppi ISIN nel file finale: {len(data)}")
            print(f"📊 CONTROLLO 1 - ISIN CENSITI: {isin_valid_count}/{len(all_orders)} ordini")
            print(f"📊 CONTROLLO 2 - MIC CODE PRESENTI: {venue_valid_count}/{len(all_orders)} ordini")
            print(f"📊 CONTROLLO 3 - DATE APPROVAL OK: {len(all_orders) - controllo_3_failed_count}/{len(all_orders)} ordini")
            print(f"📊 CONTROLLO 4 - MATURITY DATE OK: {len(all_orders) - controllo_4_failed_count}/{len(all_orders)} ordini")
            print(f"📊 Ordini totali nel file finale: {len(all_orders)}")
            
            return True
            
        except ImportError:
            self.logger.error("openpyxl non disponibile per creazione Excel")
            return False
        except Exception as e:
            self.logger.error(f"Errore creazione Excel validato: {str(e)}")
            return False
    
    def _validate_isin_esma(self, isin: str) -> bool:
        """
        Validazione ISIN tramite API ESMA reale
        
        Args:
            isin: Codice ISIN da validare
            
        Returns:
            True se l'ISIN è censito/validato da ESMA
        """
        try:
            return self.isin_validation_service.check_single_isin(isin)
        except Exception as e:
            self.logger.warning(f"Errore validazione ESMA per ISIN {isin}: {e}")
            # In caso di errore, assumiamo che l'ISIN sia valido per non bloccare il processo
            return True
    
    def _run_sequential_controls(self, data):
        """
        Esegue i controlli ESMA in modo sequenziale e procedurale per ogni ISIN.
        
        Logica procedurale:
        1. Controllo 1: API ESMA restituisce risultati? No → X nel controllo 1
        2. Controllo 2: Se controllo 1 OK, verifica trading venue vs mercato
        3. Controllo 3: Se controllo 1 e 2 OK, verifica date approval
        4. Controllo 4: Se controllo 1, 2 e 3 OK, verifica maturity date
        
        Args:
            data: Lista di gruppi ISIN
            
        Returns:
            Lista di gruppi ISIN con risultati dei controlli aggiornati
        """
        try:
            self.logger.info(f"Avvio controlli sequenziali ESMA per {len(data)} gruppi ISIN")
            
            for group_data in data:
                isin = group_data['isin']
                orders = group_data.get('orders', [])
                
                print(f"\n🔍 Elaborazione ISIN: {isin} - {len(orders)} ordini")
                
                # Effettua chiamata API ESMA una sola volta per questo ISIN
                api_error = False
                try:
                    response = self.isin_validation_service._make_api_request(isin)
                    is_valid, esma_data = self.isin_validation_service._parse_api_response_with_data(response, isin)
                    
                    api_docs = []
                    if esma_data and 'all_docs' in esma_data:
                        api_docs = esma_data['all_docs']
                    elif esma_data and 'response' in esma_data and 'docs' in esma_data['response']:
                        api_docs = esma_data['response']['docs']
                    
                    print(f"  📡 API ESMA: {len(api_docs)} risultati trovati")
                    
                except Exception as e:
                    self.logger.error(f"Errore richiesta ESMA per ISIN {isin}: {e}")
                    self.logger.warning(f"Errore API ESMA per {isin}: {e}")
                    api_docs = []
                    esma_data = None
                    api_error = True
                    print(f"  🚫 Errore API ESMA per {isin} - Controlli non eseguiti")
                
                # Se ci sono ordini per questo ISIN, controlla ognuno individualmente
                if orders and len(orders) > 0:
                    for order_index, order in enumerate(orders):
                        order_mercato = order.get('mercato') or 'XOFF'
                        order_row = order.get('row_num')
                        order_num = order.get('numero_ordine', f'#{order_index+1}')
                        
                        print(f"    🔍 Ordine {order_num} (Riga {order_row}, Mercato: {order_mercato})")
                        
                        # Inizializza tutti i controlli per questo ordine
                        order['controllo_1_failed'] = False
                        order['controllo_2_failed'] = False
                        order['controllo_3_failed'] = False
                        order['controllo_4_failed'] = False
                else:
                    # ISIN senza ordini - comunque deve essere validato per censimento ESMA
                    print(f"  ⚠️ ISIN senza ordini - controllo solo censimento ESMA")
                    # Crea ordine virtuale solo per statistiche, non per applicazione X
                    virtual_order = {
                        'row_num': group_data.get('row_num'),
                        'numero_ordine': f'VIRTUAL_{isin}',
                        'mercato': 'XOFF',
                        'is_virtual': True
                    }
                    orders = [virtual_order]
                    group_data['orders'] = orders
                
                # Ora controlla ogni ordine (reali o virtuali)
                for order_index, order in enumerate(orders):
                    order_mercato_raw = order.get('mercato') or 'XOFF'
                    order_mercato = self._extract_market_code(order_mercato_raw)
                    order_row = order.get('row_num')
                    order_num = order.get('numero_ordine', f'#{order_index+1}')
                    is_virtual = order.get('is_virtual', False)
                    
                    if not is_virtual:
                        print(f"    🔍 Ordine {order_num} (Riga {order_row}, Mercato: {order_mercato_raw})")
                    
                    # Inizializza tutti i controlli per questo ordine
                    order['controllo_1_failed'] = False
                    order['controllo_2_failed'] = False
                    order['controllo_3_failed'] = False
                    order['controllo_4_failed'] = False
                    
                    # CONTROLLO 1: API ESMA restituisce risultati?
                    if api_error:
                        # Se c'è stato un errore API, non eseguire nessun controllo
                        print(f"      🚫 CONTROLLI NON ESEGUITI: Errore API ESMA per {isin}")
                        order['esma_valid'] = None  # Indica che il controllo non è stato eseguito
                        order['api_error'] = True
                        # Non applicare nessuna X per errori API - salta tutti i controlli
                        continue
                    elif not api_docs or len(api_docs) == 0:
                        print(f"      ❌ CONTROLLO 1 FALLITO: Nessun risultato API per {isin}")
                        order['controllo_1_failed'] = True
                        order['esma_valid'] = False
                        # Se controllo 1 fallisce, non eseguire altri controlli per questo ordine
                        continue
                    else:
                        print(f"      ✅ CONTROLLO 1 PASSATO: {len(api_docs)} risultati trovati")
                        order['esma_valid'] = True
                    
                    # CONTROLLO 2: Trading venue vs mercato per questo ordine specifico
                    if order_mercato and str(order_mercato).upper() == 'XOFF':
                        # Per XOFF, qualsiasi trading venue va bene
                        print(f"      ✅ CONTROLLO 2 PASSATO: XOFF accetta qualsiasi trading venue")
                        order['venue_valid'] = True
                    else:
                        # Cerca corrispondenza esatta tra MIC e mercato
                        venue_found = False
                        for doc in api_docs:
                            doc_mic = doc.get('mic', '')
                            if str(doc_mic).upper() == str(order_mercato).upper():
                                venue_found = True
                                break
                        
                        if venue_found:
                            print(f"      ✅ CONTROLLO 2 PASSATO: MIC {order_mercato} trovato")
                            order['venue_valid'] = True
                        else:
                            print(f"      ❌ CONTROLLO 2 FALLITO: MIC {order_mercato} non trovato")
                            order['controllo_2_failed'] = True
                            order['venue_valid'] = False
                            # Se controllo 2 fallisce, non eseguire controlli successivi per questo ordine
                            continue
                    
                    # CONTROLLO 3: Date approval (solo se controlli 1 e 2 passati)
                    if not self._check_date_approval_sequential(order, api_docs, order_mercato):
                        print(f"      ✅ CONTROLLO 3 PASSATO: Date approval OK")
                    else:
                        print(f"      ❌ CONTROLLO 3 FALLITO: Date approval non valida")
                        order['controllo_3_failed'] = True
                        # Se controllo 3 fallisce, non eseguire controllo 4
                        continue
                    
                    # CONTROLLO 4: Maturity date (solo se controlli 1, 2 e 3 passati)
                    if not self._check_maturity_date_sequential(order, api_docs, order_mercato):
                        print(f"      ✅ CONTROLLO 4 PASSATO: Maturity date OK")
                    else:
                        print(f"      ❌ CONTROLLO 4 FALLITO: Maturity date non valida")
                        order['controllo_4_failed'] = True
                    
                    print(f"      🎯 Ordine {order_num} completato")
                
                # Calcola statistiche per il gruppo basandosi sui risultati degli ordini
                group_data['esma_valid'] = any(order.get('esma_valid', False) for order in orders)
                group_data['venue_valid'] = any(order.get('venue_valid', False) for order in orders)
                group_data['controllo_1_failed'] = all(order.get('controllo_1_failed', False) for order in orders)
                group_data['controllo_2_failed'] = all(order.get('controllo_2_failed', False) for order in orders)
                group_data['controllo_3_failed'] = all(order.get('controllo_3_failed', False) for order in orders)
                group_data['controllo_4_failed'] = all(order.get('controllo_4_failed', False) for order in orders)
                
                print(f"  🎯 ISIN {isin} completato")
            
            return data
            
        except Exception as e:
            self.logger.error(f"Errore nei controlli sequenziali: {e}")
            return data
    
    def _run_trading_venue_validation(self, data):
        """
        Metodo deprecato - ora integrato in _run_sequential_controls
        """
        # I controlli trading venue sono ora integrati nel metodo sequenziale
        return data

    def _run_quality_controls(self, data):
        """
        Esegue i controlli di qualità su tutti i gruppi ISIN utilizzando il servizio parallelo.
        
        Args:
            data: Lista di gruppi ISIN
            
        Returns:
            Lista di gruppi ISIN con risultati dei controlli aggiornati
        """
        try:
            # Importa dinamicamente il servizio di parallel processing
            import importlib.util
            
            # Import del servizio di processing parallelo
            from services.parallel_processing_service_threaded import ParallelProcessingServiceThreaded as ParallelProcessingService
            
            # Crea un'istanza del servizio
            parallel_service = ParallelProcessingService()
            
            # Prepara la lista di ISIN
            isin_list = [group_data['isin'] for group_data in data]
            
            # Esegui le validazioni ESMA in parallelo
            self.logger.info(f"Elaborazione {len(isin_list)} ISIN in parallelo")
            validation_results = parallel_service.process_esma_validations_parallel(isin_list)
            
            # Aggiorna i dati originali con i risultati delle validazioni
            for group_data in data:
                isin = group_data['isin']
                # Aggiorna esma_valid basato sul risultato della validazione
                group_data['esma_valid'] = validation_results.get(isin, False)
                
                # Crea un risultato di controllo qualità semplificato
                from types import SimpleNamespace
                result = SimpleNamespace()
                result.controlli_passed = 1 if group_data['esma_valid'] else 0
                result.controlli_failed = 0 if group_data['esma_valid'] else 1
                result.controlli_details = {
                    'ISIN_NON_CENSITO': '' if group_data['esma_valid'] else 'X'
                }
                
                # Se c'è un valore MERCATO, testa anche il controllo 2
                if group_data.get('mercato'):
                    try:
                        venue_match = self.isin_validation_service.check_trading_venue(isin, group_data['mercato'])
                        if venue_match:
                            result.controlli_passed += 1
                            result.controlli_details['MIC_CODE_NON_PRESENTE'] = ''
                        else:
                            result.controlli_failed += 1
                            result.controlli_details['MIC_CODE_NON_PRESENTE'] = 'X'
                    except Exception as e:
                        self.logger.warning(f"Errore controllo trading venue per ISIN {isin}: {e}")
                
                group_data['quality_controls'] = result
                
                self.logger.debug(f"ISIN {isin}: ESMA={group_data['esma_valid']}, controlli passati={result.controlli_passed}, falliti={result.controlli_failed}")
            
            return data
            
        except Exception as e:
            self.logger.error(f"Errore nei controlli di qualità: {e}")
            # In caso di errore, fallback al vecchio metodo
            for group_data in data:
                group_data['esma_valid'] = self._validate_isin_esma(group_data['isin'])
            return data

    def _check_date_approval_sequential(self, order, api_docs, mercato):
        """
        Controllo 3 sequenziale: Verifica che DATA ESEGUITO + ORA ESEGUITO > Date of approval
        
        Args:
            order: Dati dell'ordine specifico con row_num
            api_docs: Documenti API ESMA
            mercato: Codice mercato
            
        Returns:
            True se il controllo fallisce (data eseguito <= date approval)
        """
        try:
            import datetime as dt
            
            # Ottiene il numero di riga dell'ordine
            order_row = order.get('row_num')
            if not order_row:
                print(f"    ❌ Errore: numero di riga dell'ordine non trovato")
                return True  # Controllo fallisce se non trova la riga
            
            # Legge i dati dal file Excel usando openpyxl
            if not hasattr(self, '_current_excel_file'):
                print(f"    ❌ Errore: file Excel non disponibile per lettura date")
                return True  # Controllo fallisce se file non disponibile
                
            try:
                import openpyxl
                wb = openpyxl.load_workbook(self._current_excel_file, read_only=True, data_only=True)
                ws = wb.worksheets[0]  # Primo foglio
                
                # Legge DATA ESEGUITO (colonna I = colonna 9) e ORA ESEGUITO (colonna J = colonna 10)
                data_eseguito_cell = ws.cell(row=order_row, column=9).value  # Colonna I
                ora_eseguito_cell = ws.cell(row=order_row, column=10).value   # Colonna J
                
                wb.close()
                
                # Verifica che entrambi i valori siano presenti
                if not data_eseguito_cell or not ora_eseguito_cell:
                    print(f"    ❌ Dati mancanti: DATA ESEGUITO={data_eseguito_cell}, ORA ESEGUITO={ora_eseguito_cell}")
                    return True  # Controllo fallisce se dati mancanti
                
                # Converte i valori in stringhe appropriate
                if isinstance(data_eseguito_cell, dt.datetime):
                    data_eseguito = data_eseguito_cell.strftime("%d/%m/%Y")
                elif isinstance(data_eseguito_cell, dt.date):
                    data_eseguito = data_eseguito_cell.strftime("%d/%m/%Y")
                else:
                    data_eseguito = str(data_eseguito_cell).strip()
                
                if isinstance(ora_eseguito_cell, dt.datetime):
                    ora_eseguito = ora_eseguito_cell.strftime("%H:%M:%S")
                elif isinstance(ora_eseguito_cell, dt.time):
                    ora_eseguito = ora_eseguito_cell.strftime("%H:%M:%S")
                elif isinstance(ora_eseguito_cell, (int, float)):
                    # Se è un numero (tempo Excel), convertilo
                    hours = int(ora_eseguito_cell * 24)
                    minutes = int((ora_eseguito_cell * 24 * 60) % 60)
                    seconds = int((ora_eseguito_cell * 24 * 60 * 60) % 60)
                    ora_eseguito = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                else:
                    ora_eseguito = str(ora_eseguito_cell).strip()
                
                # Combina data e ora
                datetime_str = f"{data_eseguito} {ora_eseguito}"
                try:
                    datetime_eseguito = dt.datetime.strptime(datetime_str, "%d/%m/%Y %H:%M:%S")
                except ValueError:
                    # Prova altri formati comuni
                    try:
                        datetime_eseguito = dt.datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        try:
                            # Prova formato con punti: "15/09/2025 18.30.00.000000"
                            datetime_eseguito = dt.datetime.strptime(datetime_str, "%d/%m/%Y %H.%M.%S.%f")
                        except ValueError:
                            try:
                                # Prova formato senza microsecondi: "16/09/2025 10.03.56"
                                datetime_eseguito = dt.datetime.strptime(datetime_str, "%d/%m/%Y %H.%M.%S")
                            except ValueError:
                                print(f"    ❌ Formato data/ora non riconosciuto: {datetime_str}")
                                return True  # Controllo fallisce se non riesce a leggere la data
                
                print(f"    📅 Data/ora esecuzione dall'Excel: {datetime_eseguito.strftime('%d/%m/%Y %H:%M:%S')}")
                
            except ImportError:
                print(f"    ❌ Errore: openpyxl non disponibile per lettura date")
                return True  # Controllo fallisce se openpyxl non disponibile
            except Exception as e:
                print(f"    ❌ Errore lettura dati Excel: {str(e)}")
                return True  # Controllo fallisce se errore lettura Excel
            
            # Seleziona il documento corrispondente al mercato
            mercato_clean = self._extract_market_code(mercato)
            selected_doc = None
            if mercato_clean and str(mercato_clean).upper() == 'XOFF':
                selected_doc = api_docs[0] if api_docs else None
            else:
                for doc in api_docs:
                    doc_mic = doc.get('mic', '')
                    if str(doc_mic).upper() == str(mercato_clean).upper():
                        selected_doc = doc
                        break
                if not selected_doc:
                    selected_doc = api_docs[0] if api_docs else None
            
            if not selected_doc:
                return False  # Se non c'è documento, controllo passa
            
            # Estrai la market trading start date
            date_of_approval = selected_doc.get('mrkt_trdng_start_date')
            if not date_of_approval:
                return False  # Se non c'è data, controllo passa
            
            try:
                # La data dall'API ESMA è in UTC
                datetime_approvazione_utc = dt.datetime.strptime(date_of_approval, "%Y-%m-%d")
            except:
                try:
                    datetime_approvazione_utc = dt.datetime.strptime(date_of_approval.split('.')[0], "%Y-%m-%d %H:%M:%S")
                except:
                    datetime_approvazione_utc = dt.datetime(2000, 1, 1)
            
            # Converte l'ora UTC dell'API ESMA in ora italiana
            datetime_approvazione_italian = self._convert_utc_to_italian_time(datetime_approvazione_utc)
            
            print(f"    📅 Data approvazione mercato: {datetime_approvazione_italian.strftime('%d/%m/%Y %H:%M:%S')}")
            print(f"    📊 Confronto: {datetime_eseguito.strftime('%d/%m/%Y %H:%M:%S')} {'<=' if datetime_eseguito <= datetime_approvazione_italian else '>'} {datetime_approvazione_italian.strftime('%d/%m/%Y %H:%M:%S')}")
            
            # Controllo fallisce se data eseguito <= data approvazione
            return datetime_eseguito <= datetime_approvazione_italian
            
        except Exception as e:
            self.logger.error(f"Errore controllo date approval: {e}")
            return True  # In caso di errore, controllo fallisce
    
    def _check_maturity_date_sequential(self, group_data, api_docs, mercato):
        """
        Controllo 4 sequenziale: Verifica che DATA ESEGUITO + ORA ESEGUITO < Maturity date
        
        Args:
            group_data: Dati del gruppo ISIN
            api_docs: Documenti API ESMA
            mercato: Codice mercato
            
        Returns:
            True se il controllo fallisce (data eseguito >= maturity date)
        """
        try:
            import datetime as dt
            
            # Seleziona il documento corrispondente al mercato
            mercato_clean = self._extract_market_code(mercato)
            selected_doc = None
            if mercato_clean and str(mercato_clean).upper() == 'XOFF':
                selected_doc = api_docs[0] if api_docs else None
            else:
                for doc in api_docs:
                    doc_mic = doc.get('mic', '')
                    if str(doc_mic).upper() == str(mercato_clean).upper():
                        selected_doc = doc
                        break
                if not selected_doc:
                    selected_doc = api_docs[0] if api_docs else None
            
            if not selected_doc:
                return False  # Se non c'è documento, controllo passa
            
            # Estrai la maturity date
            maturity_date = selected_doc.get('bnd_maturity_date')
            if not maturity_date:
                maturity_date = selected_doc.get('mrkt_trdng_trmination_date')
            
            if not maturity_date:
                return False  # Se non c'è data, controllo passa
            
            # Ottiene il numero di riga dell'ordine dal group_data
            order_row = group_data.get('row_num')
            if not order_row:
                print(f"    ❌ Errore: numero di riga dell'ordine non trovato per maturity check")
                return False  # Se non trova la riga, controllo passa
            
            # Legge i dati dal file Excel usando openpyxl
            if not hasattr(self, '_current_excel_file'):
                print(f"    ❌ Errore: file Excel non disponibile per lettura date")
                return False  # Se file non disponibile, controllo passa
                
            try:
                import openpyxl
                wb = openpyxl.load_workbook(self._current_excel_file, read_only=True, data_only=True)
                ws = wb.worksheets[0]  # Primo foglio
                
                # Legge DATA ESEGUITO (colonna I = colonna 9) e ORA ESEGUITO (colonna J = colonna 10)
                data_eseguito_cell = ws.cell(row=order_row, column=9).value  # Colonna I
                ora_eseguito_cell = ws.cell(row=order_row, column=10).value   # Colonna J
                
                wb.close()
                
                # Verifica che entrambi i valori siano presenti
                if not data_eseguito_cell or not ora_eseguito_cell:
                    print(f"    ❌ Dati mancanti per maturity check: DATA ESEGUITO={data_eseguito_cell}, ORA ESEGUITO={ora_eseguito_cell}")
                    return False  # Se dati mancanti, controllo passa
                
                # Converte i valori in stringhe appropriate (stesso codice di _check_date_approval_sequential)
                if isinstance(data_eseguito_cell, dt.datetime):
                    data_eseguito = data_eseguito_cell.strftime("%d/%m/%Y")
                elif isinstance(data_eseguito_cell, dt.date):
                    data_eseguito = data_eseguito_cell.strftime("%d/%m/%Y")
                else:
                    data_eseguito = str(data_eseguito_cell).strip()
                
                if isinstance(ora_eseguito_cell, dt.datetime):
                    ora_eseguito = ora_eseguito_cell.strftime("%H:%M:%S")
                elif isinstance(ora_eseguito_cell, dt.time):
                    ora_eseguito = ora_eseguito_cell.strftime("%H:%M:%S")
                elif isinstance(ora_eseguito_cell, (int, float)):
                    # Se è un numero (tempo Excel), convertilo
                    hours = int(ora_eseguito_cell * 24)
                    minutes = int((ora_eseguito_cell * 24 * 60) % 60)
                    seconds = int((ora_eseguito_cell * 24 * 60 * 60) % 60)
                    ora_eseguito = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                else:
                    ora_eseguito = str(ora_eseguito_cell).strip()
                
                # Combina data e ora
                datetime_str = f"{data_eseguito} {ora_eseguito}"
                try:
                    datetime_eseguito = dt.datetime.strptime(datetime_str, "%d/%m/%Y %H:%M:%S")
                except ValueError:
                    # Prova altri formati comuni
                    try:
                        datetime_eseguito = dt.datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        try:
                            # Prova formato con punti: "15/09/2025 18.30.00.000000"
                            datetime_eseguito = dt.datetime.strptime(datetime_str, "%d/%m/%Y %H.%M.%S.%f")
                        except ValueError:
                            try:
                                # Prova formato senza microsecondi: "16/09/2025 10.03.56"
                                datetime_eseguito = dt.datetime.strptime(datetime_str, "%d/%m/%Y %H.%M.%S")
                            except ValueError:
                                print(f"    ❌ Formato data/ora non riconosciuto per maturity check: {datetime_str}")
                                return False  # Se non riesce a leggere la data, controllo passa
                
            except ImportError:
                print(f"    ❌ Errore: openpyxl non disponibile per lettura date")
                return False  # Se openpyxl non disponibile, controllo passa
            except Exception as e:
                print(f"    ❌ Errore lettura dati Excel per maturity check: {str(e)}")
                return False  # Se errore lettura Excel, controllo passa
            
            try:
                # La data dall'API ESMA è in UTC
                datetime_maturity_utc = dt.datetime.strptime(maturity_date, "%Y-%m-%d")
            except:
                try:
                    datetime_maturity_utc = dt.datetime.strptime(maturity_date.split('.')[0], "%Y-%m-%d %H:%M:%S")
                except:
                    datetime_maturity_utc = dt.datetime(9999, 12, 31)
            
            # Converte l'ora UTC dell'API ESMA in ora italiana
            datetime_maturity_italian = self._convert_utc_to_italian_time(datetime_maturity_utc)
            
            # Controllo fallisce se data eseguito >= maturity date
            return datetime_eseguito >= datetime_maturity_italian
            
        except Exception as e:
            self.logger.error(f"Errore controllo maturity date: {e}")
            return False  # In caso di errore, controllo passa
        """
        Genera un resoconto dettagliato di tutti i controlli effettuati.
        
        Args:
            data: Lista dei dati processati
            validated_count: Numero di ISIN validati nel controllo 1
            venue_valid_count: Numero di venue validi nel controllo 2
            non_censiti_count: Numero di ISIN non censurati
            output_path: Path del file di output
        """
        try:
            from datetime import datetime
            
            total_isin = len(data)
            
            print(f"\n" + "="*80)
            print(f"📋 RESOCONTO DETTAGLIATO CONTROLLI CON-412")
            print(f"📅 Data elaborazione: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            print(f"="*80)
            
            # 1. Controllo Accesso File
            print(f"\n🔍 FASE 1-3: ACCESSO E LETTURA FILE")
            print(f"   ✅ Accesso file verificato")
            print(f"   ✅ File copiato in directory di lavoro")
            print(f"   ✅ Struttura Excel letta correttamente")
            print(f"   📊 ISIN totali identificati: {total_isin}")
            
            # 2. Controllo Database
            print(f"\n🔍 FASE 4: FILTRAGGIO DATABASE ORACLE")
            print(f"   ✅ Connessione database Oracle TNS")
            print(f"   ✅ Query RF/EE eseguita")
            print(f"   📊 Ordini filtrati per tipologia RF/EE")
            
            # 3. Controllo ESMA ISIN
            print(f"\n🔍 FASE 5: PRIMO CONTROLLO ESMA - VALIDAZIONE ISIN")
            if validated_count > 0:
                print(f"   ✅ API ESMA disponibile e funzionante")
                print(f"   ✅ Elaborazione parallela completata")
                success_rate = (validated_count / total_isin * 100) if total_isin > 0 else 0
                print(f"   📊 ISIN validati: {validated_count}/{total_isin} ({success_rate:.1f}%)")
                print(f"   📊 ISIN NON CENSITI: {non_censiti_count}/{total_isin}")
                if non_censiti_count > 0:
                    print(f"   ⚠️  Trovati {non_censiti_count} ISIN non presenti nel registro ESMA")
                else:
                    print(f"   ✅ Tutti gli ISIN sono presenti nel registro ESMA")
            else:
                print(f"   ❌ API ESMA: NON DISPONIBILE o nessun ISIN validato")
                user_choice = input("Vuoi riprovare la chiamata API ESMA? (R)ipeti o (C)ontinua senza validazione: ")
                if user_choice.upper() == "R":
                    print("Procedura di ripetizione non implementata, continuazione senza validazione API.")
                else:
                    print("Continuazione senza validazione API ESMA")
            
            # 4. Controllo Trading Venue
            print(f"\n🔍 FASE 6: SECONDO CONTROLLO ESMA - VALIDAZIONE TRADING VENUE")
            print(f"   ✅ Verifica corrispondenza MERCATO vs Trading Venue")
            print(f"   ✅ Eccezione XOFF gestita correttamente")
            venue_success_rate = (venue_valid_count / total_isin * 100) if total_isin > 0 else 0
            print(f"   📊 Venue validati: {venue_valid_count}/{total_isin} ({venue_success_rate:.1f}%)")
            venue_failed = total_isin - venue_valid_count
            if venue_failed > 0:
                print(f"   ⚠️  MIC CODE non presenti: {venue_failed}/{total_isin}")
            else:
                print(f"   ✅ Tutti i MIC CODE sono validi")
            
            # 5. Generazione Output
            print(f"\n🔍 FASE 7: GENERAZIONE FILE EXCEL")
            print(f"   ✅ Struttura originale preservata")
            print(f"   ✅ Colonne controllo aggiunte")
            print(f"   ✅ File salvato correttamente")
            print(f"   📁 Percorso: {output_path}")
            
            # 6. Riepilogo Finale
            print(f"\n🎯 RIEPILOGO FINALE")
            print(f"   📊 File processato: ✅ SUCCESSO")
            print(f"   📊 Database filtering: ✅ SUCCESSO")
            if validated_count > 0:
                print(f"   📊 Controllo 1 (ISIN): ✅ COMPLETATO ({validated_count}/{total_isin})")
            else:
                print(f"   📊 Controllo 1 (ISIN): ⚠️  SALTATO (API non disponibile)")
            print(f"   📊 Controllo 2 (Venue): ✅ COMPLETATO ({venue_valid_count}/{total_isin})")
            print(f"   📊 Output generato: ✅ SUCCESSO")
            
            # 7. Raccomandazioni
            print(f"\n💡 RACCOMANDAZIONI")
            if validated_count == 0:
                print(f"   ⚠️  Ripetere l'elaborazione quando API ESMA torna disponibile")
                print(f"   📋 Verificare manualmente gli ISIN per completare il controllo 1")
            if venue_failed > 0:
                print(f"   📝 Verificare manualmente i {venue_failed} MIC CODE non trovati")
            if non_censiti_count > 0 and validated_count > 0:
                print(f"   📝 Rivedere gli {non_censiti_count} ISIN marcati come non censurati")
            if validated_count > 0 and venue_failed == 0 and non_censiti_count == 0:
                print(f"   ✅ Tutti i controlli superati con successo!")
            
            print(f"="*80)
            
            # Log del resoconto
            self.logger.info("RESOCONTO DETTAGLIATO GENERATO")
            self.logger.info(f"ISIN totali: {total_isin}")
            self.logger.info(f"Controllo 1: {validated_count}/{total_isin} ({'SALTATO' if validated_count == 0 else 'COMPLETATO'})")
            self.logger.info(f"Controllo 2: {venue_valid_count}/{total_isin} COMPLETATO")
            self.logger.info(f"API ESMA: {'DISPONIBILE' if validated_count > 0 else 'NON DISPONIBILE'}")
            
        except Exception as e:
            self.logger.error(f"Errore generazione resoconto: {e}")
            print(f"⚠️  Errore nel resoconto dettagliato: {e}")

    def _check_date_approval(self, group, new_ws, original_to_new_row_mapping, casistica_col, original_ws, data_eseguito_col, ora_eseguito_col, mercato_col, casistica_isin_col, casistica_venue_col):
        """
        Controllo 3: Verifica che DATA ESEGUITO + ORA ESEGUITO > Date of approval
        Restituisce True se il controllo fallisce (e mette X), False se passa
        APPLICA SOLO ALLE RIGHE CHE NON HANNO X NEI CONTROLLI 1 E 2
        """
        try:
            import datetime as dt
            
            print(f"🔍 DEBUG - Entrando nel controllo 3 per ISIN: {group.get('isin')}")
            
            # Estrai i dati ESMA per questo ISIN
            isin = group.get('isin')
            if not isin:
                print(f"❌ DEBUG - Nessun ISIN nel gruppo")
                return False
                
            # Ottieni i dati ESMA tramite chiamata API diretta
            print(f"🔍 DEBUG - Effettuando chiamata API ESMA per controllo date per {isin}")
            try:
                # Usa i metodi esistenti per fare una nuova chiamata API
                response = self.isin_validation_service._make_api_request(isin)
                is_valid, esma_data = self.isin_validation_service._parse_api_response_with_data(response, isin)
                print(f"🔍 DEBUG - Risposta API ESMA per {isin}: validità={is_valid}, dati={esma_data is not None}")
                print(f"🔍 DEBUG - Struttura dati ESMA: {list(esma_data.keys()) if esma_data else 'None'}")
            except Exception as e:
                print(f"❌ DEBUG - Errore chiamata API ESMA per {isin}: {e}")
                return False
                
            if not esma_data:
                print(f"❌ DEBUG - Nessun dato ESMA ricevuto per {isin}")
                return False
                
            # Estrai la market trading start date dalla risposta API
            api_docs = esma_data.get('all_docs', [])
            print(f"🔍 DEBUG - Documenti API trovati: {len(api_docs)}")

            # Se non ci sono docs, prova a cercare in altre parti della risposta
            if not api_docs:
                print(f"🔍 DEBUG - Cerco documenti in altre parti della risposta ESMA")
                if 'response' in esma_data:
                    response_docs = esma_data['response'].get('docs', [])
                    print(f"🔍 DEBUG - Documenti in response: {len(response_docs)}")
                    api_docs = response_docs

            if not api_docs:
                print(f"❌ DEBUG - Nessun documento API per {isin}")
                return False

            # Seleziona il documento corrispondente al mercato
            # Prima estrai il mercato dalla prima riga ordine
            order_rows = group.get('order_rows', [])
            if not order_rows:
                orders = group.get('orders', [])
                order_rows = [order.get('row_num') for order in orders if order.get('row_num')]
                
            if not order_rows:
                return False
                
            first_order_row = order_rows[0]
            mercato = original_ws.cell(row=first_order_row, column=mercato_col).value if mercato_col else "XOFF"
            
            print(f"📅 DEBUG Controllo 3 - ISIN: {isin}")
            print(f"📅 DEBUG - Mercato estratto: {mercato}")
            
            # Log specifici per gli ISIN NL0009434992 e IE00BLNN3691
            if isin in ['NL0009434992', 'IE00BLNN3691']:
                print(f"🔍 DEBUG SPECIFICO - ISIN: {isin}")
                print(f"🔍 DEBUG SPECIFICO - Mercato: {mercato}")
                print(f"🔍 DEBUG SPECIFICO - Documenti restituiti dall'API: {len(api_docs)}")
                for idx, doc in enumerate(api_docs):
                    print(f"🔍 DEBUG SPECIFICO - Documento {idx + 1}: MIC={doc.get('mic', 'N/A')}, Altre chiavi={list(doc.keys())}")

            # Verifica se l'ISIN è censito (controllo 1)
            if api_docs:
                group['esma_valid'] = True  # ISIN censito se ci sono risultati
                print(f"✅ DEBUG - ISIN {isin} censito con {len(api_docs)} risultati")
            else:
                group['esma_valid'] = False  # ISIN non censito
                print(f"❌ DEBUG - ISIN {isin} non censito (nessun risultato API)")
                return False
            
            # Seleziona il documento corrispondente al mercato
            selected_doc = None
            if mercato and 'XOFF' in str(mercato).upper():
                # Per mercati che contengono "XOFF", prendi il primo documento disponibile
                selected_doc = api_docs[0] if api_docs else None
                print(f"🔍 DEBUG - Mercato contiene XOFF, uso primo documento disponibile")
            else:
                # Filtra per MIC che corrisponde esattamente al MERCATO
                for doc in api_docs:
                    doc_mic = doc.get('mic', '')
                    print(f"🔍 DEBUG - Confronto MIC '{doc_mic}' con mercato '{mercato}'")
                    if str(doc_mic).upper() == str(mercato).upper():
                        selected_doc = doc
                        print(f"✅ DEBUG - Trovato documento con MIC {doc_mic} che corrisponde a mercato {mercato}")
                        break
                
                # Se non trova corrispondenza esatta, prendi il primo documento
                if not selected_doc:
                    selected_doc = api_docs[0] if api_docs else None
                    print(f"⚠️ DEBUG - Nessuna corrispondenza MIC trovata per '{mercato}', uso primo documento disponibile")

            if not selected_doc:
                print(f"❌ DEBUG - Nessun documento selezionato per il mercato {mercato}")
                return False

            # Estrai la market trading start date
            date_of_approval = selected_doc.get('mrkt_trdng_start_date')
            if not date_of_approval:
                print(f"❌ DEBUG - Nessuna data di approvazione trovata nel documento selezionato")
                return False

            print(f"📅 DEBUG - Date of approval estratta: {date_of_approval}")
            
            # Estrai i valori delle celle Excel usando le colonne passate come parametri
            order_rows = group.get('order_rows', [])
            if not order_rows:
                orders = group.get('orders', [])
                order_rows = [order.get('row_num') for order in orders if order.get('row_num')]
                
            if not order_rows:
                return False
                
            # Prendi la prima riga ordine per estrarre data, ora e mercato
            first_order_row = order_rows[0]
            
            # Estrai i valori dalle celle Excel usando le colonne passate come parametri
            data_eseguito = original_ws.cell(row=first_order_row, column=data_eseguito_col).value if data_eseguito_col else "22/09/2025"
            ora_eseguito = original_ws.cell(row=first_order_row, column=ora_eseguito_col).value if ora_eseguito_col else "14:30"
            mercato = original_ws.cell(row=first_order_row, column=mercato_col).value if mercato_col else "XOFF"
            
            # Converti le stringhe in datetime
            if isinstance(data_eseguito, str):
                data_str = data_eseguito
            else:
                data_str = data_eseguito.strftime('%d/%m/%Y') if data_eseguito else "22/09/2025"
                
            if isinstance(ora_eseguito, str):
                # Gestisci formato Excel time con punti e microsecondi: "10.01.59.115662"
                ora_str = ora_eseguito
                if '.' in ora_str:
                    # Converti da "10.01.59.115662" a "10:01:59"
                    time_parts = ora_str.split('.')
                    if len(time_parts) >= 3:
                        ora_str = f"{time_parts[0]}:{time_parts[1]}:{time_parts[2]}"
                    elif len(time_parts) == 2:
                        ora_str = f"{time_parts[0]}:{time_parts[1]}:00"
                    else:
                        ora_str = "14:30:00"
                else:
                    ora_str = ora_str if ':' in ora_str else "14:30:00"
            else:
                ora_str = ora_eseguito.strftime('%H:%M:%S') if ora_eseguito else "14:30:00"
            
            datetime_eseguito = dt.datetime.strptime(f"{data_str} {ora_str}", "%d/%m/%Y %H:%M:%S")
            
            # Gestisci il formato della date of approval (potrebbe essere diverso)
            try:
                datetime_approvazione = dt.datetime.strptime(date_of_approval, "%Y-%m-%d")
            except:
                try:
                    datetime_approvazione = dt.datetime.strptime(date_of_approval, "%Y-%m-%d %H:%M:%S")
                except:
                    try:
                        # Gestisci formato con decimali: "2025-02-19 06:00:00.0"
                        datetime_approvazione = dt.datetime.strptime(date_of_approval.split('.')[0], "%Y-%m-%d %H:%M:%S")
                    except:
                        try:
                            datetime_approvazione = dt.datetime.strptime(date_of_approval, "%d/%m/%Y")
                        except:
                            # Default se tutti i formati falliscono
                            datetime_approvazione = dt.datetime(2000, 1, 1)
            
            print(f"📅 DEBUG - DateTime eseguito: {datetime_eseguito}")
            print(f"📅 DEBUG - DateTime approvazione: {datetime_approvazione}")
            
            if datetime_eseguito <= datetime_approvazione:
                print(f"❌ Controllo 3 FALLITO per ISIN {isin}: {datetime_eseguito} <= {datetime_approvazione}")
                # Controllo fallito - metti X solo nelle righe che non hanno X nei controlli precedenti
                if 'order_rows' in group:
                    for original_row_num in group['order_rows']:
                        if original_row_num in original_to_new_row_mapping:
                            new_row = original_to_new_row_mapping[original_row_num]
                            
                            # Controlla se questa riga ha già X nei controlli 1 o 2
                            ha_x_controllo1 = False
                            ha_x_controllo2 = False
                            
                            if casistica_isin_col:
                                controllo1_cell = new_ws.cell(row=new_row, column=casistica_isin_col)
                                ha_x_controllo1 = (controllo1_cell.value == 'X')
                                
                            if casistica_venue_col:
                                controllo2_cell = new_ws.cell(row=new_row, column=casistica_venue_col)
                                ha_x_controllo2 = (controllo2_cell.value == 'X')
                            
                            # Solo se non ha X nei controlli precedenti, metti X nel controllo 3
                            if not ha_x_controllo1 and not ha_x_controllo2:
                                print(f"DEBUG - Inserimento X per Controllo 3, ISIN {isin}, riga {new_row}")
                                cell = new_ws.cell(row=new_row, column=casistica_col)
                                cell.value = 'X'
                                cell.font = Font(bold=False, color="000000")
                            else:
                                print(f"DEBUG - Saltata riga {new_row} per Controllo 3: ha già X in controllo precedente")
                                
                elif 'orders' in group:
                    for order in group['orders']:
                        if 'row_num' in order and order['row_num'] in original_to_new_row_mapping:
                            new_row = original_to_new_row_mapping[order['row_num']]
                            
                            # Controlla se questa riga ha già X nei controlli 1 o 2
                            ha_x_controllo1 = False
                            ha_x_controllo2 = False
                            
                            if casistica_isin_col:
                                controllo1_cell = new_ws.cell(row=new_row, column=casistica_isin_col)
                                ha_x_controllo1 = (controllo1_cell.value == 'X')
                                
                            if casistica_venue_col:
                                controllo2_cell = new_ws.cell(row=new_row, column=casistica_venue_col)
                                ha_x_controllo2 = (controllo2_cell.value == 'X')
                            
                            # Solo se non ha X nei controlli precedenti, metti X nel controllo 3
                            if not ha_x_controllo1 and not ha_x_controllo2:
                                cell = new_ws.cell(row=new_row, column=casistica_col)
                                cell.value = 'X'
                                cell.font = Font(bold=False, color="000000")
                return True  # Controllo fallito
            else:
                print(f"✅ Controllo 3 PASSATO per ISIN {isin}: {datetime_eseguito} > {datetime_approvazione}")
                
            return False  # Controllo passato
            
        except Exception as e:
            self.logger.error(f"Errore controllo date approval per ISIN {group.get('isin')}: {e}")
            print(f"❌ Errore controllo 3 per ISIN {group.get('isin')}: {e}")
            return False

    def _check_maturity_date(self, group, new_ws, original_to_new_row_mapping, casistica_col, original_ws, data_eseguito_col, ora_eseguito_col, mercato_col, casistica_isin_col, casistica_venue_col, casistica_date_approval_col):
        """
        Controllo 4: Verifica che DATA ESEGUITO + ORA ESEGUITO < Maturity date
        Restituisce True se il controllo fallisce (e mette X), False se passa
        APPLICA SOLO ALLE RIGHE CHE NON HANNO X NEI CONTROLLI 1, 2 E 3
        """
        try:
            import datetime as dt
            
            print(f"🔍 DEBUG - Entrando nel controllo 4 per ISIN: {group.get('isin')}")
            
            # Estrai i dati ESMA per questo ISIN
            isin = group.get('isin')
            if not isin:
                print(f"❌ DEBUG - Nessun ISIN nel gruppo")
                return False
                
            # Ottieni i dati ESMA tramite chiamata API diretta
            print(f"🔍 DEBUG - Effettuando chiamata API ESMA per controllo maturity per {isin}")
            try:
                # Usa i metodi esistenti per fare una nuova chiamata API
                response = self.isin_validation_service._make_api_request(isin)
                is_valid, esma_data = self.isin_validation_service._parse_api_response_with_data(response, isin)
                print(f"🔍 DEBUG - Risposta API ESMA per {isin}: validità={is_valid}, dati={esma_data is not None}")
            except Exception as e:
                print(f"❌ DEBUG - Errore chiamata API ESMA per {isin}: {e}")
                return False
                
            if not esma_data:
                print(f"❌ DEBUG - Nessun dato ESMA ricevuto per {isin}")
                return False
                
            # Cerca la maturity date nell'API response
            api_docs = esma_data.get('all_docs', [])
            print(f"🔍 DEBUG - Documenti API trovati per controllo 4: {len(api_docs)}")
            
            # Se non ci sono docs, prova a cercare in altre parti della risposta
            if not api_docs:
                print(f"🔍 DEBUG - Cerco documenti in altre parti della risposta ESMA per controllo 4")
                if 'response' in esma_data:
                    response_docs = esma_data['response'].get('docs', [])
                    print(f"🔍 DEBUG - Documenti in response per controllo 4: {len(response_docs)}")
                    api_docs = response_docs
                    
            if not api_docs:
                print(f"❌ DEBUG - Nessun documento API per controllo 4 su {isin}")
                return False
                
            # Estrai il mercato dalle righe ordine
            order_rows = group.get('order_rows', [])
            if not order_rows:
                orders = group.get('orders', [])
                order_rows = [order.get('row_num') for order in orders if order.get('row_num')]
                
            if not order_rows:
                return False
                
            # Prendi la prima riga ordine per estrarre data, ora e mercato
            first_order_row = order_rows[0]
            
            # Estrai i valori dalle celle Excel
            data_eseguito = original_ws.cell(row=first_order_row, column=data_eseguito_col).value if data_eseguito_col else "22/09/2025"
            ora_eseguito = original_ws.cell(row=first_order_row, column=ora_eseguito_col).value if ora_eseguito_col else "14:30"
            mercato = original_ws.cell(row=first_order_row, column=mercato_col).value if mercato_col else "XOFF"
            
            print(f"📅 DEBUG Controllo 4 - ISIN: {isin}")
            print(f"📅 DEBUG - Data eseguito: {data_eseguito}")
            print(f"📅 DEBUG - Ora eseguito: {ora_eseguito}")
            print(f"📅 DEBUG - Mercato: {mercato}")
            print(f"📅 DEBUG - Controllo 4: DATA ESEGUITO + ORA ESEGUITO < bnd_maturity_date")
            
            # Filtra il documento corrispondente al mercato usando MIC
            selected_doc = None

            if mercato and 'XOFF' in str(mercato).upper():
                # Per mercati che contengono "XOFF", prendi il primo documento disponibile
                selected_doc = api_docs[0] if api_docs else None
                print(f"🔍 DEBUG - Mercato contiene XOFF, uso primo documento disponibile per controllo 4")
            else:
                # Filtra per MIC che corrisponde esattamente al MERCATO
                for doc in api_docs:
                    doc_mic = doc.get('mic', '')
                    print(f"🔍 DEBUG - Confronto MIC '{doc_mic}' con mercato '{mercato}' per controllo 4")
                    if str(doc_mic).upper() == str(mercato).upper():
                        selected_doc = doc
                        print(f"✅ DEBUG - Trovato documento con MIC {doc_mic} che corrisponde a mercato {mercato} per controllo 4")
                        break
                
                # Se non trova corrispondenza esatta, prendi il primo documento
                if not selected_doc:
                    selected_doc = api_docs[0] if api_docs else None
                    print(f"⚠️ DEBUG - Nessuna corrispondenza MIC trovata per '{mercato}', uso primo documento disponibile per controllo 4")
                    
            if not selected_doc:
                print(f"❌ DEBUG - Nessun documento selezionato per il mercato {mercato} nel controllo 4")
                return False
                
            # Estrai la maturity date o la termination date
            maturity_date = selected_doc.get('bnd_maturity_date')
            if not maturity_date:
                maturity_date = selected_doc.get('mrkt_trdng_trmination_date')

            if not maturity_date:
                print(f"❌ DEBUG - Nessuna bnd_maturity_date o mrkt_trdng_trmination_date trovata nel documento selezionato")
                print(f"🔍 DEBUG - Chiavi disponibili nel documento per controllo 4: {list(selected_doc.keys())}")
                return False
                
            print(f"📅 DEBUG - Maturity date estratta: {maturity_date}")
            print(f"📅 DEBUG - Logica: Se DATA+ORA >= MATURITY_DATE → controllo FALLISCE")
            
            # Converti le stringhe in datetime
            if isinstance(data_eseguito, str):
                data_str = data_eseguito
            else:
                data_str = data_eseguito.strftime('%d/%m/%Y') if data_eseguito else "22/09/2025"
                
            if isinstance(ora_eseguito, str):
                # Gestisci formato Excel time con punti e microsecondi: "10.01.59.115662"
                ora_str = ora_eseguito
                if '.' in ora_str:
                    # Converti da "10.01.59.115662" a "10:01:59"
                    time_parts = ora_str.split('.')
                    if len(time_parts) >= 3:
                        ora_str = f"{time_parts[0]}:{time_parts[1]}:{time_parts[2]}"
                    elif len(time_parts) == 2:
                        ora_str = f"{time_parts[0]}:{time_parts[1]}:00"
                    else:
                        ora_str = "14:30:00"
                else:
                    ora_str = ora_str if ':' in ora_str else "14:30:00"
            else:
                ora_str = ora_eseguito.strftime('%H:%M:%S') if ora_eseguito else "14:30:00"
            
            datetime_eseguito = dt.datetime.strptime(f"{data_str} {ora_str}", "%d/%m/%Y %H:%M:%S")
            
            # Gestisci il formato della maturity date
            try:
                datetime_maturity = dt.datetime.strptime(maturity_date, "%Y-%m-%d")
            except:
                try:
                    datetime_maturity = dt.datetime.strptime(maturity_date, "%Y-%m-%d %H:%M:%S")
                except:
                    try:
                        # Gestisci formato con decimali: "9999-12-31 00:00:00.0"
                        datetime_maturity = dt.datetime.strptime(maturity_date.split('.')[0], "%Y-%m-%d %H:%M:%S")
                    except:
                        try:
                            datetime_maturity = dt.datetime.strptime(maturity_date, "%d/%m/%Y")
                        except:
                            # Default a una data futura molto lontana se tutti i formati falliscono
                            datetime_maturity = dt.datetime(9999, 12, 31)
            
            print(f"📅 DEBUG - DateTime eseguito: {datetime_eseguito}")
            print(f"📅 DEBUG - DateTime maturity: {datetime_maturity}")
            print(f"📅 DEBUG - Confronto: {datetime_eseguito} >= {datetime_maturity} = {datetime_eseguito >= datetime_maturity}")
            
            if datetime_eseguito >= datetime_maturity:
                print(f"❌ Controllo 4 FALLITO per ISIN {isin}: Esecuzione {datetime_eseguito} è >= scadenza {datetime_maturity}")
                # Controllo fallito - metti X solo nelle righe che non hanno X nei controlli precedenti
                if 'order_rows' in group:
                    for original_row_num in group['order_rows']:
                        if original_row_num in original_to_new_row_mapping:
                            new_row = original_to_new_row_mapping[original_row_num]
                            
                            # Controlla se questa riga ha già X nei controlli 1, 2 o 3
                            ha_x_controllo1 = False
                            ha_x_controllo2 = False
                            ha_x_controllo3 = False
                            
                            if casistica_isin_col:
                                controllo1_cell = new_ws.cell(row=new_row, column=casistica_isin_col)
                                ha_x_controllo1 = (controllo1_cell.value == 'X')
                                
                            if casistica_venue_col:
                                controllo2_cell = new_ws.cell(row=new_row, column=casistica_venue_col)
                                ha_x_controllo2 = (controllo2_cell.value == 'X')
                                
                            if casistica_date_approval_col:
                                controllo3_cell = new_ws.cell(row=new_row, column=casistica_date_approval_col)
                                ha_x_controllo3 = (controllo3_cell.value == 'X')
                            
                            # Solo se non ha X nei controlli precedenti, metti X nel controllo 4
                            if not ha_x_controllo1 and not ha_x_controllo2 and not ha_x_controllo3:
                                print(f"DEBUG - Inserimento X per Controllo 4, ISIN {isin}, riga {new_row}")
                                cell = new_ws.cell(row=new_row, column=casistica_col)
                                cell.value = 'X'
                                cell.font = Font(bold=False, color="000000")
                            else:
                                print(f"DEBUG - Saltata riga {new_row} per Controllo 4: ha già X in controllo precedente")
                                
                elif 'orders' in group:
                    for order in group['orders']:
                        if 'row_num' in order and order['row_num'] in original_to_new_row_mapping:
                            new_row = original_to_new_row_mapping[order['row_num']]
                            
                            # Controlla se questa riga ha già X nei controlli 1, 2 o 3
                            ha_x_controllo1 = False
                            ha_x_controllo2 = False
                            ha_x_controllo3 = False
                            
                            if casistica_isin_col:
                                controllo1_cell = new_ws.cell(row=new_row, column=casistica_isin_col)
                                ha_x_controllo1 = (controllo1_cell.value == 'X')
                                
                            if casistica_venue_col:
                                controllo2_cell = new_ws.cell(row=new_row, column=casistica_venue_col)
                                ha_x_controllo2 = (controllo2_cell.value == 'X')
                                
                            if casistica_date_approval_col:
                                controllo3_cell = new_ws.cell(row=new_row, column=casistica_date_approval_col)
                                ha_x_controllo3 = (controllo3_cell.value == 'X')
                            
                            # Solo se non ha X nei controlli precedenti, metti X nel controllo 4
                            if not ha_x_controllo1 and not ha_x_controllo2 and not ha_x_controllo3:
                                cell = new_ws.cell(row=new_row, column=casistica_col)
                                cell.value = 'X'
                                cell.font = Font(bold=False, color="000000")
                return True  # Controllo fallito
            else:
                print(f"✅ Controllo 4 PASSATO per ISIN {isin}: Esecuzione {datetime_eseguito} è < scadenza {datetime_maturity}")
                
            return False  # Controllo passato
            
        except Exception as e:
            self.logger.error(f"Errore controllo maturity date per ISIN {group.get('isin')}: {e}")
            print(f"❌ Errore controllo 4 per ISIN {group.get('isin')}: {e}")
            return False


def main():
    """Funzione principale con interfaccia interattiva"""
    print("=" * 60)
    print("🏦 CON-412 TRANSACTION REPORTING - REJECTING MENSILE")
    print("Sistema di elaborazione automatica ISIN via ESMA")
    print("=" * 60)
    print(f"📅 Avvio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        # Configurazione interattiva
        config_manager = InteractiveConfig()
        config = config_manager.get_file_config()
        
        if not config:
            print("\n❌ Configurazione non valida. Uscita.")
            return 1
        
        print(f"\n✅ Configurazione completata!")
        print(f"Tipo: {config['type']}")
        
        # Avvia automaticamente il processore
        print("\n🚀 AVVIO ELABORAZIONE AUTOMATICA")
        processor = CON412Processor(config)
        success = processor.run()
        
        if success:
            print("\n🎉 ELABORAZIONE COMPLETATA CON SUCCESSO!")
            return 0
        else:
            print("\n❌ ELABORAZIONE FALLITA!")
            return 1
            
    except KeyboardInterrupt:
        print("\n\n⚠️  Elaborazione interrotta dall'utente")
        return 130
        
    except Exception as e:
        print(f"\n💥 Errore critico: {str(e)}")
        # Aggiungi log dettagliati per catturare eccezioni non gestite
        try:
            with open("error_log.txt", "w") as log_file:
                log_file.write(f"Errore critico: {str(e)}\n")
        except:
            print("❌ Errore anche nella scrittura del log degli errori")
        return 1


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
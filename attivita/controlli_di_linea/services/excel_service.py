"""
Servizio per la gestione e creazione di file Excel.
Gestisce la creazione di file Excel ottimizzati con supporto per grandi dataset.
"""

import pandas as pd
from typing import List, Dict, Any, Optional, Tuple, Iterator
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from models.config import Config, OutputConfig
from models.component import LookupResult, ProcessingStats
from utils.file_utils import ensure_directory, get_available_disk_space


class ExcelService:
    """Servizio per la gestione dei file Excel."""
    
    def __init__(self, config: Config):
        """
        Inizializza il servizio Excel.
        
        Args:
            config: Configurazione dell'applicazione
        """
        self.config = config
        self.logger = logging.getLogger(__name__)
    
    def create_excel_files(self, results: Iterator[LookupResult], 
                          output_config: OutputConfig,
                          stats: Optional[ProcessingStats] = None) -> List[str]:
        """
        Crea file Excel dai risultati del lookup.
        
        Args:
            results: Iterator dei risultati del lookup
            output_config: Configurazione di output
            stats: Statistiche di elaborazione (opzionale)
            
        Returns:
            Lista dei file creati
        """
        ensure_directory(output_config.output_directory)
        
        # Calcola il numero massimo di righe dinamicamente
        max_rows = self._calculate_dynamic_max_rows()
        
        created_files = []
        current_file_data = []
        current_row_count = 0
        part_number = 1
        
        self.logger.info(f"Creazione file Excel con massimo {max_rows} righe per file")
        
        # Processa tutti i risultati
        for result in results:
            current_file_data.append(result.to_dict())
            current_row_count += 1
            
            # Se raggiungiamo il limite, salva il file
            if current_row_count >= max_rows:
                file_path = self._save_excel_part(
                    current_file_data, 
                    output_config, 
                    part_number
                )
                if file_path:
                    created_files.append(file_path)
                
                # Reset per il prossimo file
                current_file_data = []
                current_row_count = 0
                part_number += 1
        
        # Salva i dati rimanenti se presenti
        if current_file_data:
            file_path = self._save_excel_part(
                current_file_data, 
                output_config, 
                part_number
            )
            if file_path:
                created_files.append(file_path)
        
        # Crea file di statistiche se richiesto
        if stats and self.config.create_summary_report:
            summary_file = self._create_summary_file(stats, output_config, created_files)
            if summary_file:
                created_files.append(summary_file)
        
        self.logger.info(f"Creati {len(created_files)} file Excel")
        return created_files
    
    def _calculate_dynamic_max_rows(self) -> int:
        """
        Calcola dinamicamente il numero massimo di righe per file.
        
        Returns:
            Numero massimo di righe
        """
        # Limite teorico di Excel
        excel_limit = 1048576
        
        # Considera la memoria disponibile
        try:
            import psutil
            available_memory_gb = psutil.virtual_memory().available / (1024**3)
            
            # Stima conservativa: ~200 bytes per riga per i nostri dati
            estimated_rows_per_gb = 5000000  # ~5M righe per GB
            memory_based_limit = int(available_memory_gb * estimated_rows_per_gb * 0.5)  # Usa solo 50% della RAM
            
            # Prendi il minimo tra limite Excel e limite memoria
            calculated_limit = min(excel_limit - 1000, memory_based_limit)  # -1000 per header e margine
            
        except ImportError:
            # Fallback se psutil non è disponibile
            calculated_limit = excel_limit - 1000
        
        # Non scendere sotto un minimo ragionevole
        return max(100000, min(calculated_limit, self.config.max_rows_per_file))
    
    def _save_excel_part(self, data: List[Dict[str, Any]], 
                        output_config: OutputConfig, 
                        part_number: int) -> Optional[str]:
        """
        Salva una parte dei dati in un file Excel.
        
        Args:
            data: Dati da salvare
            output_config: Configurazione di output
            part_number: Numero della parte
            
        Returns:
            Percorso del file creato
        """
        if not data:
            return None
        
        try:
            # Genera nome file
            file_path = output_config.get_output_filename(
                part_number if len(data) > 0 else None
            )
            
            # Crea DataFrame
            df = pd.DataFrame(data)
            
            # Salva usando il motore configurato
            if self.config.excel_engine == "xlsxwriter":
                self._save_with_xlsxwriter(df, file_path)
            else:
                self._save_with_openpyxl(df, file_path)
            
            self.logger.info(f"Salvato file Excel: {file_path} ({len(data)} righe)")
            return file_path
            
        except Exception as e:
            self.logger.error(f"Errore nel salvataggio Excel parte {part_number}: {e}")
            return None
    
    def _save_with_xlsxwriter(self, df: pd.DataFrame, file_path: str):
        """
        Salva un DataFrame usando XlsxWriter (ottimizzato per performance).
        
        Args:
            df: DataFrame da salvare
            file_path: Percorso del file
        """
        with pd.ExcelWriter(file_path, engine='xlsxwriter', 
                           options={'strings_to_numbers': False}) as writer:
            
            df.to_excel(writer, sheet_name='Risultati', index=False)
            
            # Ottieni il workbook e worksheet per la formattazione
            workbook = writer.book
            worksheet = writer.sheets['Risultati']
            
            # Formattazione header
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D7E4BC',
                'border': 1
            })
            
            # Applica formattazione header
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Auto-fit colonne
            for i, col in enumerate(df.columns):
                # Calcola larghezza ottimale
                max_len = max(
                    df[col].astype(str).map(len).max(),  # Max lunghezza dati
                    len(str(col))  # Lunghezza header
                )
                # Limita la larghezza
                worksheet.set_column(i, i, min(max_len + 2, 50))
            
            # Freeze prima riga
            worksheet.freeze_panes(1, 0)
    
    def _save_with_openpyxl(self, df: pd.DataFrame, file_path: str):
        """
        Salva un DataFrame usando openpyxl (più flessibile per formattazione).
        
        Args:
            df: DataFrame da salvare
            file_path: Percorso del file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Risultati"
        
        # Inserisci i dati
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Formattazione header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-fit colonne
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze prima riga
        ws.freeze_panes = 'A2'
        
        wb.save(file_path)
    
    def _create_summary_file(self, stats: ProcessingStats, 
                           output_config: OutputConfig,
                           data_files: List[str]) -> Optional[str]:
        """
        Crea un file di riepilogo con statistiche.
        
        Args:
            stats: Statistiche di elaborazione
            output_config: Configurazione di output
            data_files: Lista dei file di dati creati
            
        Returns:
            Percorso del file di riepilogo
        """
        try:
            summary_path = output_config.get_output_filename(extension="xlsx").replace(
                ".xlsx", "_RIEPILOGO.xlsx"
            )
            
            wb = Workbook()
            
            # Sheet statistiche
            ws_stats = wb.active
            ws_stats.title = "Statistiche"
            
            # Header
            ws_stats['A1'] = "RIEPILOGO ELABORAZIONE COMPONENTI ORACLE"
            ws_stats['A1'].font = Font(bold=True, size=14)
            
            # Dati statistiche
            stats_data = [
                ["Data Elaborazione", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["", ""],
                ["TOTALI", ""],
                ["Tabelle Elaborate", stats.total_tables],
                ["Tabelle Trovate", stats.matched_tables],
                ["Tabelle Non Trovate", stats.unmatched_tables],
                ["Errori", stats.errors],
                ["", ""],
                ["PERFORMANCE", ""],
                ["Percentuale Successo", f"{stats.match_rate:.1f}%"],
                ["Tempo Elaborazione", f"{stats.duration:.2f} secondi" if stats.duration else "N/A"],
                ["Velocità", f"{stats.total_tables / stats.duration:.1f} tabelle/sec" if stats.duration and stats.duration > 0 else "N/A"],
                ["", ""],
                ["FILE GENERATI", ""],
                ["Numero File", len(data_files)]
            ]
            
            for i, (label, value) in enumerate(stats_data, start=3):
                ws_stats[f'A{i}'] = label
                ws_stats[f'B{i}'] = value
                
                if label and not value:  # Headers
                    ws_stats[f'A{i}'].font = Font(bold=True)
            
            # Sheet file generati
            if data_files:
                ws_files = wb.create_sheet("File Generati")
                ws_files['A1'] = "File"
                ws_files['B1'] = "Dimensione (MB)"
                ws_files['C1'] = "Righe Stimate"
                
                for i, file_path in enumerate(data_files, start=2):
                    ws_files[f'A{i}'] = os.path.basename(file_path)
                    
                    try:
                        size_mb = os.path.getsize(file_path) / (1024 * 1024)
                        ws_files[f'B{i}'] = f"{size_mb:.2f}"
                    except:
                        ws_files[f'B{i}'] = "N/A"
                    
                    # Stima righe (molto approssimativa)
                    try:
                        estimated_rows = int(size_mb * 10000)  # Stima ~10k righe per MB
                        ws_files[f'C{i}'] = estimated_rows
                    except:
                        ws_files[f'C{i}'] = "N/A"
                
                # Formattazione header
                for cell in ws_files[1]:
                    cell.font = Font(bold=True)
            
            wb.save(summary_path)
            self.logger.info(f"Creato file riepilogo: {summary_path}")
            return summary_path
            
        except Exception as e:
            self.logger.error(f"Errore nella creazione del file riepilogo: {e}")
            return None
    
    def validate_excel_compatibility(self, estimated_rows: int) -> Tuple[bool, List[str]]:
        """
        Valida la compatibilità con Excel per il numero di righe stimato.
        
        Args:
            estimated_rows: Numero stimato di righe
            
        Returns:
            Tupla (compatibile, lista_avvisi)
        """
        warnings = []
        compatible = True
        
        # Limite Excel
        excel_limit = 1048576
        if estimated_rows > excel_limit:
            files_needed = (estimated_rows // excel_limit) + 1
            warnings.append(f"Dataset troppo grande per un singolo file Excel. "
                          f"Saranno creati {files_needed} file.")
        
        # Verifica spazio disco
        try:
            output_dir = self.config.output_directory
            available_space = get_available_disk_space(output_dir)
            
            # Stima dimensione file (molto approssimativa)
            estimated_size_mb = estimated_rows * 0.0002  # ~200 bytes per riga
            estimated_size_bytes = estimated_size_mb * 1024 * 1024
            
            if estimated_size_bytes > available_space:
                warnings.append(f"Spazio disco insufficiente. "
                              f"Richiesti ~{estimated_size_mb:.0f}MB, "
                              f"disponibili {available_space/(1024*1024):.0f}MB")
                compatible = False
                
        except Exception:
            warnings.append("Impossibile verificare lo spazio disco disponibile")
        
        return compatible, warnings
    
    def optimize_dataframe_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Ottimizza un DataFrame per l'esportazione Excel.
        
        Args:
            df: DataFrame da ottimizzare
            
        Returns:
            DataFrame ottimizzato
        """
        optimized_df = df.copy()
        
        # Converti tutto in stringhe per evitare problemi di tipo
        for col in optimized_df.columns:
            optimized_df[col] = optimized_df[col].astype(str)
        
        # Sostituisci valori NaN/None con stringhe vuote
        optimized_df = optimized_df.fillna('')
        
        # Rimuovi caratteri problematici
        for col in optimized_df.select_dtypes(include=['object']).columns:
            optimized_df[col] = optimized_df[col].str.replace('\x00', '', regex=False)  # Null characters
            optimized_df[col] = optimized_df[col].str.replace('\r\n', ' ', regex=False)  # Line breaks
            optimized_df[col] = optimized_df[col].str.replace('\n', ' ', regex=False)
            optimized_df[col] = optimized_df[col].str.replace('\r', ' ', regex=False)
        
        # Limita la lunghezza delle celle (Excel ha un limite di ~32k caratteri)
        for col in optimized_df.columns:
            optimized_df[col] = optimized_df[col].astype(str).str[:32000]
        
        return optimized_df
    
    def create_excel_from_dataframe(self, df: pd.DataFrame, 
                                  output_path: str,
                                  sheet_name: str = "Dati") -> bool:
        """
        Crea un file Excel da un DataFrame.
        
        Args:
            df: DataFrame da salvare
            output_path: Percorso del file di output
            sheet_name: Nome del foglio
            
        Returns:
            True se il salvataggio è riuscito
        """
        try:
            # Ottimizza il DataFrame
            optimized_df = self.optimize_dataframe_for_excel(df)
            
            # Salva
            if self.config.excel_engine == "xlsxwriter":
                self._save_with_xlsxwriter(optimized_df, output_path)
            else:
                self._save_with_openpyxl(optimized_df, output_path)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Errore nel salvataggio Excel {output_path}: {e}")
            return False
